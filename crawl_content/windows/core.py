import re
import time
import base64
import requests
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from io import BytesIO
from PIL import Image
import logging

# Cấu hình logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def setup_driver():
    """Khởi tạo Chrome driver với cấu hình tối ưu"""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    options.page_load_strategy = 'eager'  # Không đợi tải hết trang

    return webdriver.Chrome(options=options)


def download_image(url, save_path, timeout=15):
    """Tải ảnh từ URL về máy"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://diakov.net/'
        }

        logger.info(f"  📥 Đang tải ảnh: {url[:60]}...")
        response = requests.get(url, headers=headers, timeout=timeout, stream=True)
        response.raise_for_status()

        # Kiểm tra content-type
        content_type = response.headers.get('content-type', '')
        if 'image' not in content_type:
            logger.warning(f"  ⚠️ Không phải ảnh: {content_type}")
            return False

        img = Image.open(BytesIO(response.content))

        # Convert sang RGB nếu cần
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        # Resize nếu ảnh quá lớn
        if img.width > 800 or img.height > 800:
            img.thumbnail((800, 800), Image.Resampling.LANCZOS)

        img.save(save_path, 'JPEG', quality=85)
        logger.info(f"  ✓ Đã lưu: {save_path.name}")
        return True

    except Exception as e:
        logger.error(f"  ✗ Không tải được ảnh {url[:60]}: {str(e)}")
        return False


def extract_download_links(driver):
    """Trích xuất các link download từ div.zvcwers"""
    download_links = []

    try:
        # Tìm div chứa link download
        download_div = driver.find_element(By.CSS_SELECTOR, "div.zvcwers")

        # Lấy tất cả các thẻ <a> trong div
        links = download_div.find_elements(By.TAG_NAME, "a")

        for link in links:
            href = link.get_attribute("href")
            text = link.text.strip()

            # Bỏ qua link VIP hoặc link rỗng
            if not href or not text or "buy.php" in href or "VIP" in text:
                continue

            # Lấy tên host từ text hoặc từ URL
            if text:
                host_name = text
            else:
                # Parse domain từ URL
                match = re.search(r'https?://(?:www\.)?([^/]+)', href)
                host_name = match.group(1) if match else "Unknown"

            download_links.append({
                "host": host_name,
                "url": href
            })

        logger.info(f"  → Tìm thấy {len(download_links)} link download")

    except Exception as e:
        logger.warning(f"  → Không tìm thấy link download: {str(e)}")

    return download_links


def extract_article_data(driver, url):
    """Trích xuất dữ liệu từ một bài viết"""
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 10)

        # Đợi nội dung chính load
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1.ftitle.card-title")))

        # 1. Lấy tag và title
        title_elem = driver.find_element(By.CSS_SELECTOR, "h1.ftitle.card-title")
        tag = title_elem.find_element(By.TAG_NAME, "a").text.strip()
        title = title_elem.find_element(By.TAG_NAME, "u").text.strip()

        # 2. Lấy mô tả
        body_elem = driver.find_element(By.CSS_SELECTOR, "div.card-body.sh")
        full_text = body_elem.text.strip()

        # Loại bỏ phần không cần thiết
        patterns = [
            r'Tính năng.*',
            r'Đặc điểm.*',
            r'HĐH:.*',
            r'Tải xuống.*',
            r'từ.*Github.*',
            r'Đã cảm ơn:.*'
        ]

        description = full_text
        for pattern in patterns:
            description = re.split(pattern, description, flags=re.IGNORECASE)[0]

        # Giới hạn độ dài mô tả
        words = description.split()[:150]
        description = ' '.join(words).strip()

        # 3. Lấy ảnh (tối đa 3 ảnh)
        # Scroll để lazy load images
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        imgs = driver.find_elements(By.CSS_SELECTOR, "div.card-body.sh img")
        image_urls = []

        logger.info(f"  → Tìm thấy {len(imgs)} thẻ img")

        for idx, img in enumerate(imgs, start=1):
            # Thử nhiều attribute
            src = (img.get_attribute("src") or
                   img.get_attribute("data-src") or
                   img.get_attribute("data-lazy-src"))

            logger.debug(f"    Ảnh {idx}: src={src}")

            # Bỏ qua ảnh placeholder gif
            if not src or 'data:image/gif' in src or src.endswith('.gif'):
                continue

            # Xử lý URL tương đối
            if src.startswith("/"):
                src = "https://diakov.net" + src

            # Chỉ lấy ảnh thực sự (không phải icon, spoiler, rating)
            if src.startswith("http"):
                # Kiểm tra kích thước ảnh để loại bỏ icon nhỏ
                try:
                    width = img.get_attribute("width") or img.size.get('width', 0)
                    height = img.get_attribute("height") or img.size.get('height', 0)

                    # Chỉ lấy ảnh có kích thước hợp lý
                    if width and height and (int(width) < 50 or int(height) < 50):
                        continue
                except:
                    pass

                # Bỏ qua icon, spoiler, rating
                if not any(x in src.lower() for x in ['icon', 'spoiler', 'rating', 'dleimages']):
                    image_urls.append(src)
                    logger.info(f"    ✓ Lấy ảnh: {src[:80]}...")

                    if len(image_urls) >= 3:
                        break

        # 4. Lấy link download
        download_links = extract_download_links(driver)

        logger.info(f"✓ Crawled: {title}")

        return {
            "url": url,
            "tag": tag,
            "title": title,
            "description": description,
            "images": image_urls,
            "downloads": download_links
        }

    except Exception as e:
        logger.error(f"✗ Lỗi crawl {url}: {str(e)}")
        return None


def save_to_excel(data, output_file="soft_data.xlsx"):
    """Lưu dữ liệu vào Excel với format đẹp"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Software Data"

    # Header
    headers = ["STT", "Tag", "Title", "Description", "Image 1 URL", "Image 2 URL", "Image 3 URL", "Download Links",
               "URL"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 40

    row_num = 2
    for idx, item in enumerate(data, start=1):
        if not item:
            continue

        # Thêm dữ liệu text
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=item["tag"])
        ws.cell(row=row_num, column=3, value=item["title"])
        ws.cell(row=row_num, column=4, value=item["description"])

        # Thêm URL ảnh vào 3 cột riêng biệt
        images = item.get("images", [])
        for img_idx in range(3):
            col = 5 + img_idx  # Cột E, F, G
            if img_idx < len(images):
                ws.cell(row=row_num, column=col, value=images[img_idx])
            else:
                ws.cell(row=row_num, column=col, value="")

        # Format download links
        download_text = ""
        for dl in item.get("downloads", []):
            download_text += f"{dl['host']}: {dl['url']}\n"
        ws.cell(row=row_num, column=8, value=download_text.strip())

        ws.cell(row=row_num, column=9, value=item["url"])

        # Wrap text cho description và download links
        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_num, column=8).alignment = Alignment(wrap_text=True, vertical='top')

        # Wrap text cho image URLs
        for col in range(5, 8):
            ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        row_num += 1

    wb.save(output_file)
    logger.info(f"✅ Đã lưu dữ liệu vào '{output_file}'")


def crawl_core(urls):
    """Hàm chính để crawl danh sách URLs"""
    driver = setup_driver()
    data = []

    try:
        total = len(urls)
        for idx, url in enumerate(urls, start=1):
            logger.info(f"[{idx}/{total}] Đang crawl: {url}")

            article_data = extract_article_data(driver, url)
            if article_data:
                data.append(article_data)

            # Nghỉ ngắn giữa các request
            time.sleep(1)

    finally:
        driver.quit()

    return data

# Module này export các hàm để sử dụng ở nơi khác:
# - setup_driver(): Khởi tạo Chrome driver
# - extract_article_data(driver, url): Crawl một bài viết
# - crawl_core(urls): Crawl danh sách URLs
# - save_to_excel(data, output_file): Lưu dữ liệu ra Excel